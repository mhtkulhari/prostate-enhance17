import argparse
import csv
import json
import logging
import os
import random
import sys

import numpy as np
import torch
import torch.backends.cudnn as cudnn
from torch.utils.data import DataLoader
from torchvision import transforms

from networks.lesion_models import build_segmentation_model
from dataloaders.dataloader import FundusSegmentation, ProstateSegmentation, ProstateLesionSegmentation, MNMSSegmentation
import dataloaders.custom_transforms as tr
from utils import losses, metrics, ramps, util
from medpy.metric import binary

parser = argparse.ArgumentParser()
parser.add_argument('--dataset', type=str, default='prostate', choices=['fundus', 'prostate', 'prostate_lesion', 'MNMS'])
parser.add_argument("--save_name", type=str, default="debug", help="experiment_name")
parser.add_argument("--overwrite", action='store_true')
parser.add_argument(
    "--model",
    type=str,
    default="unet",
    choices=["unet", "fpn", "deeplabv3plus", "segformer", "unetpp", "attunet", "resunet", "scseunet"],
    help="model_name"
)
parser.add_argument('--encoder_name', type=str, default=None,
                    help='encoder for SMP models: fpn/deeplabv3plus default resnet34, segformer default mit_b0')
parser.add_argument('--encoder_weights', type=str, default='imagenet',
                    help='encoder weights for SMP models; use None to train encoder from scratch')
parser.add_argument("--gpu", type=str, default='0')
parser.add_argument('--data_root', type=str, default='../../data/Prostate_Lesion', help='override dataset root path')
parser.add_argument('--eval',type=bool, default=True)

parser.add_argument("--test_bs", type=int, default=1)
parser.add_argument('--domain_num', type=int, default=6)
parser.add_argument('--lb_domain', type=int, default=1)

parser.add_argument('--save_img',action='store_true')
parser.add_argument('--lesion_modalities', type=str, default='t2w,adc',
                    help='comma-separated registered modality names for prostate_lesion. Default t2w,adc resolves to image_t2w,image_adc layout')
parser.add_argument('--lesion_modality_dirs', type=str, default=None)
parser.add_argument('--lesion_norm', type=str, default='minmax', choices=['legacy', 'minmax', 'zscore'])
parser.add_argument('--add_adc_sobel', type=int, default=0)
parser.add_argument('--post_min_area', type=int, default=0)
parser.add_argument('--post_topk', type=int, default=0)
parser.add_argument('--post_fill_holes', type=int, default=0)
args = parser.parse_args()


def split_csv(value):
    if value is None:
        return []
    return [item.strip() for item in str(value).split(',') if item.strip()]

def to_2d(input_tensor):
    input_tensor = input_tensor.unsqueeze(1)
    tensor_list = []
    temp_prob = input_tensor == torch.ones_like(input_tensor)
    tensor_list.append(temp_prob)
    temp_prob2 = input_tensor > torch.zeros_like(input_tensor)
    tensor_list.append(temp_prob2)
    output_tensor = torch.cat(tensor_list, dim=1)
    return output_tensor.float()

def to_3d(input_tensor):
    input_tensor = input_tensor.unsqueeze(1)
    tensor_list = []
    for i in range(1, 4):
        temp_prob = input_tensor == i * torch.ones_like(input_tensor)
        tensor_list.append(temp_prob)
    output_tensor = torch.cat(tensor_list, dim=1)
    return output_tensor.float()

if args.dataset == 'fundus':
    part = ['cup', 'disc']
    dataset = FundusSegmentation
elif args.dataset == 'prostate':
    part = ['base']
    dataset = ProstateSegmentation
elif args.dataset == 'prostate_lesion':
    part = ['lesion']
    dataset = ProstateLesionSegmentation
elif args.dataset == 'MNMS':
    part = ['lv', 'myo', 'rv']
    dataset = MNMSSegmentation
n_part = len(part)
dice_calcu = {'fundus': metrics.dice_coeff_2label, 'prostate': metrics.dice_coeff, 'prostate_lesion': metrics.dice_coeff, 'MNMS': metrics.dice_coeff_3label}


def _safe_div(numerator, denominator):
    return float(numerator) / float(denominator) if denominator else 0.0


def _binary_dice(pred_bool, gt_bool):
    intersection = int(np.logical_and(pred_bool, gt_bool).sum())
    pred_pixels = int(pred_bool.sum())
    gt_pixels = int(gt_bool.sum())
    return _safe_div(2 * intersection, pred_pixels + gt_pixels)


def _binary_jaccard(pred_bool, gt_bool):
    intersection = int(np.logical_and(pred_bool, gt_bool).sum())
    union = int(np.logical_or(pred_bool, gt_bool).sum())
    return _safe_div(intersection, union)


def _case_id_from_img_name(img_name):
    base = os.path.splitext(os.path.basename(str(img_name)))[0]
    parts = base.split('_')
    if len(parts) > 1 and parts[0].startswith('Dom'):
        base = '_'.join(parts[1:])
    return base.rsplit('_', 1)[0] if '_' in base else base


def _new_lesion_stats(domain):
    return {
        'domain': domain,
        'total_slices': 0,
        'positive_slices': 0,
        'empty_slices': 0,
        'tp': 0,
        'fp': 0,
        'tn': 0,
        'fn': 0,
        'positive_pred_empty_slices': 0,
        'empty_pred_positive_slices': 0,
        'positive_dice_sum': 0.0,
        'positive_jaccard_sum': 0.0,
        'positive_hd95_sum': 0.0,
        'positive_asd_sum': 0.0,
        'pred_pixels_on_empty_total': 0,
        'pred_pixels_on_empty_max': 0,
        'pixel_tp': 0,
        'pixel_fp': 0,
        'pixel_tn': 0,
        'pixel_fn': 0,
    }


def _new_case_stats(domain, case_id):
    return {
        'domain': domain,
        'case_id': case_id,
        'total_slices': 0,
        'positive_slices': 0,
        'empty_slices': 0,
        'pixel_tp': 0,
        'pixel_fp': 0,
        'pixel_fn': 0,
        'positive_dice_sum': 0.0,
        'positive_jaccard_sum': 0.0,
        'positive_hd95_sum': 0.0,
        'positive_asd_sum': 0.0,
        'pred_volume_slices': [],
        'gt_volume_slices': [],
    }


def _finalize_case_stats(stats):
    tp = int(stats['pixel_tp'])
    fp = int(stats['pixel_fp'])
    fn = int(stats['pixel_fn'])
    positive = int(stats['positive_slices'])
    dice_den = 2 * tp + fp + fn
    jaccard_den = tp + fp + fn

    pred_slices = stats.get('pred_volume_slices', [])
    gt_slices = stats.get('gt_volume_slices', [])
    volume_hd95 = None
    volume_asd = None

    if pred_slices and gt_slices:
        pred_volume = np.stack(pred_slices, axis=0).astype(bool)
        gt_volume = np.stack(gt_slices, axis=0).astype(bool)
        pred_volume_pos = bool(pred_volume.any())
        gt_volume_pos = bool(gt_volume.any())

        if pred_volume_pos or gt_volume_pos:
            if pred_volume_pos and gt_volume_pos:
                try:
                    volume_hd95 = float(binary.hd95(pred_volume, gt_volume))
                    volume_asd = float(binary.asd(pred_volume, gt_volume))
                except Exception:
                    volume_hd95 = 100.0
                    volume_asd = 100.0
            else:
                volume_hd95 = 100.0
                volume_asd = 100.0

    return {
        'domain': stats['domain'],
        'case_id': stats['case_id'],
        'total_slices': int(stats['total_slices']),
        'positive_slices': positive,
        'empty_slices': int(stats['empty_slices']),
        'pixel_tp': tp,
        'pixel_fp': fp,
        'pixel_fn': fn,
        'volume_dice': None if dice_den == 0 else float(2 * tp) / float(dice_den),
        'volume_jaccard': None if jaccard_den == 0 else float(tp) / float(jaccard_den),
        'volume_hd95': volume_hd95,
        'volume_asd': volume_asd,
        'positive_dice': None if positive == 0 else float(stats['positive_dice_sum']) / float(positive),
        'positive_jaccard': None if positive == 0 else float(stats['positive_jaccard_sum']) / float(positive),
        'positive_hd95': None if positive == 0 else float(stats['positive_hd95_sum']) / float(positive),
        'positive_asd': None if positive == 0 else float(stats['positive_asd_sum']) / float(positive),
        'dice_denominator': dice_den,
    }

def _finalize_lesion_stats(stats):
    total = int(stats['total_slices'])
    positive = int(stats['positive_slices'])
    empty = int(stats['empty_slices'])
    tp = int(stats['tp'])
    fp = int(stats['fp'])
    tn = int(stats['tn'])
    fn = int(stats['fn'])

    # IMPORTANT:
    # Do not compute volume Dice/Jaccard here from pooled domain/all pixels.
    # Volume metrics are calculated only case-wise in _finalize_case_stats(),
    # then averaged either over cases or equally over domains.
    return {
        'domain': stats['domain'],
        'total_slices': total,
        'positive_slices': positive,
        'empty_slices': empty,
        'tp': tp,
        'fp': fp,
        'tn': tn,
        'fn': fn,
        'positive_pred_empty_slices': int(stats['positive_pred_empty_slices']),
        'empty_pred_positive_slices': int(stats['empty_pred_positive_slices']),
        'positive_dice': _safe_div(stats['positive_dice_sum'], positive),
        'positive_jaccard': _safe_div(stats['positive_jaccard_sum'], positive),
        'positive_hd95': _safe_div(stats['positive_hd95_sum'], positive),
        'positive_asd': _safe_div(stats['positive_asd_sum'], positive),
        'slice_accuracy': _safe_div(tp + tn, total),
        'slice_sensitivity': _safe_div(tp, tp + fn),
        'slice_specificity': _safe_div(tn, tn + fp),
        'slice_precision': _safe_div(tp, tp + fp),
        'slice_npv': _safe_div(tn, tn + fn),
        'slice_f1': _safe_div(2 * tp, 2 * tp + fp + fn),
        'slice_balanced_accuracy': 0.5 * (_safe_div(tp, tp + fn) + _safe_div(tn, tn + fp)),
        'empty_false_positive_rate': _safe_div(fp, empty),
        'empty_true_negative_rate': _safe_div(tn, empty),
        'pred_pixels_on_empty_total': int(stats['pred_pixels_on_empty_total']),
        'pred_pixels_on_empty_mean': _safe_div(stats['pred_pixels_on_empty_total'], empty),
        'pred_pixels_on_empty_max': int(stats['pred_pixels_on_empty_max']),
    }


def _casewise_volume_mean(case_results):
    valid_cases = [item for item in case_results if item['volume_dice'] is not None]
    valid_hd_cases = [item for item in case_results if item.get('volume_hd95') is not None]
    valid_asd_cases = [item for item in case_results if item.get('volume_asd') is not None]
    return {
        'volume_dice': _safe_div(sum(item['volume_dice'] for item in valid_cases), len(valid_cases)),
        'volume_jaccard': _safe_div(sum(item['volume_jaccard'] for item in valid_cases), len(valid_cases)),
        'volume_hd95': _safe_div(sum(item['volume_hd95'] for item in valid_hd_cases), len(valid_hd_cases)),
        'volume_asd': _safe_div(sum(item['volume_asd'] for item in valid_asd_cases), len(valid_asd_cases)),
        'valid_cases': len(valid_cases),
        'valid_hd95_cases': len(valid_hd_cases),
        'valid_asd_cases': len(valid_asd_cases),
        'skipped_empty_empty_cases': len(case_results) - len(valid_cases),
    }


def _casewise_overall_metric_mean(case_results):
    valid_volume_cases = [item for item in case_results if item['volume_dice'] is not None]
    valid_hd_cases = [item for item in case_results if item.get('volume_hd95') is not None]
    valid_asd_cases = [item for item in case_results if item.get('volume_asd') is not None]
    return {
        'volume_dice': _safe_div(sum(item['volume_dice'] for item in valid_volume_cases), len(valid_volume_cases)),
        'volume_jaccard': _safe_div(sum(item['volume_jaccard'] for item in valid_volume_cases), len(valid_volume_cases)),
        'volume_hd95': _safe_div(sum(item['volume_hd95'] for item in valid_hd_cases), len(valid_hd_cases)),
        'volume_asd': _safe_div(sum(item['volume_asd'] for item in valid_asd_cases), len(valid_asd_cases)),
        'valid_volume_cases': len(valid_volume_cases),
        'valid_hd95_cases': len(valid_hd_cases),
        'valid_asd_cases': len(valid_asd_cases),
    }



def _pct(value):
    return round(100.0 * float(value), 2) if value is not None else ''


def _num(value):
    return round(float(value), 2) if value is not None else ''


def _confusion_text(stats):
    # Order requested in Excel: [TP, FN, FP, TN]
    return '[{}, {}, {}, {}]'.format(
        int(stats.get('tp', 0)),
        int(stats.get('fn', 0)),
        int(stats.get('fp', 0)),
        int(stats.get('tn', 0)),
    )


def _save_lesion_summary_excel(args, epoch, domain_results, domain_mean, overall_case_mean, all_slice_stats):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    headers = [
        'Model',
        'Domain',
        'DC',
        'JC',
        'HD95',
        'ASD',
        '[TP, FN, FP, TN]',
        'Sensitivity',
        'Specificity',
        'Precision',
        'F1 Score',
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'
    ws.append(headers)

    model_name = 'Our'
    rows = []

    for idx, d in enumerate(domain_results):
        dom_num = int(str(d['domain']).replace('Dom', ''))
        dom_label = '{} (L)'.format(d['domain']) if dom_num == int(args.lb_domain) else d['domain']

        rows.append([
            model_name if idx == 0 else '',
            dom_label,
            _pct(d['volume_dice']),
            _pct(d['volume_jaccard']),
            _num(d['volume_hd95']),
            _num(d['volume_asd']),
            _confusion_text(d),
            _pct(d['slice_sensitivity']),
            _pct(d['slice_specificity']),
            _pct(d['slice_precision']),
            _pct(d['slice_f1']),
        ])

    # Average row: domain-wise average of volume metrics only.
    rows.append([
        '',
        'Average',
        _pct(domain_mean['volume_dice']),
        _pct(domain_mean['volume_jaccard']),
        _num(domain_mean['volume_hd95']),
        _num(domain_mean['volume_asd']),
        '',
        '',
        '',
        '',
        '',
    ])

    # Overall row:
    # DC/JC/HD95/ASD are calculated case-wise over all cases first, then averaged.
    # Final confusion and classification metrics are from all slices combined.
    rows.append([
        '',
        'Overall',
        _pct(overall_case_mean['volume_dice']),
        _pct(overall_case_mean['volume_jaccard']),
        _num(overall_case_mean['volume_hd95']),
        _num(overall_case_mean['volume_asd']),
        _confusion_text(all_slice_stats),
        _pct(all_slice_stats['slice_sensitivity']),
        _pct(all_slice_stats['slice_specificity']),
        _pct(all_slice_stats['slice_precision']),
        _pct(all_slice_stats['slice_f1']),
    ])

    for row in rows:
        ws.append(row)

    last_row = ws.max_row
    ws.merge_cells(start_row=2, start_column=1, end_row=last_row, end_column=1)
    ws.cell(row=2, column=1).value = model_name

    header_fill = PatternFill('solid', fgColor='D9D9D9')
    body_fill = PatternFill('solid', fgColor='DDEBF7')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            cell.font = Font(name='Calibri', size=11, bold=(cell.row == 1 or cell.column == 2))
            cell.fill = header_fill if cell.row == 1 else body_fill

    ws.cell(row=2, column=1).font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')

    for col in range(1, len(headers) + 1):
        max_len = max(len(str(ws.cell(row=r, column=col).value or '')) for r in range(1, last_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = max(12, min(max_len + 3, 24))

    ws.row_dimensions[1].height = 22
    for r in range(2, last_row + 1):
        ws.row_dimensions[r].height = 22

    excel_path = os.path.join(os.getcwd(), '{}_test_metrics_epoch_{}.xlsx'.format(args.save_name, epoch))
    wb.save(excel_path)

    logging.info('saved lesion summary excel: {}'.format(excel_path))
    print('saved lesion summary excel:', excel_path)

    return excel_path

def _test_prostate_lesion_with_empty_metrics(args, model, test_dataloader, epoch):
    model.eval()
    domain_results = []
    slice_rows = []
    case_stats = {}
    overall_stats = _new_lesion_stats('FULL')

    for domain_idx, cur_dataloader in enumerate(test_dataloader):
        domain_code = domain_idx + 1
        domain_stats = _new_lesion_stats('Dom{}'.format(domain_code))

        for batch_num, sample in enumerate(cur_dataloader):
            data = sample['image'].cuda()
            mask = sample['label'].cuda().gt(0).long()
            output = model(data).cpu()
            mask = mask.cpu()

            pred_label = torch.max(torch.softmax(output, dim=1), dim=1)[1]
            if args.post_min_area > 0 or args.post_topk > 0 or args.post_fill_holes:
                pred_np = util.postprocess_binary_batch(
                    pred_label.numpy(),
                    min_area=args.post_min_area,
                    topk=args.post_topk,
                    fill_holes=bool(args.post_fill_holes),
                )
                pred_label = torch.from_numpy(pred_np).long()

            pred_onehot = pred_label.clone().unsqueeze(1)
            mask_onehot = mask.clone().unsqueeze(1)

            for j in range(len(data)):
                img_name = sample['img_name'][j]
                case_id = _case_id_from_img_name(img_name)
                case_key = 'Dom{}/{}'.format(domain_code, case_id)
                if case_key not in case_stats:
                    case_stats[case_key] = _new_case_stats('Dom{}'.format(domain_code), case_id)

                pred_bool = np.asarray(pred_onehot[j, 0], dtype=bool)
                gt_bool = np.asarray(mask_onehot[j, 0], dtype=bool)
                pred_pixels = int(pred_bool.sum())
                gt_pixels = int(gt_bool.sum())
                intersection_pixels = int(np.logical_and(pred_bool, gt_bool).sum())
                pred_pos = pred_pixels > 0
                gt_pos = gt_pixels > 0

                pixel_fp = int(np.logical_and(pred_bool, np.logical_not(gt_bool)).sum())
                pixel_tn = int(np.logical_and(np.logical_not(pred_bool), np.logical_not(gt_bool)).sum())
                pixel_fn = int(np.logical_and(np.logical_not(pred_bool), gt_bool).sum())

                for stats in (domain_stats, overall_stats):
                    stats['total_slices'] += 1
                    stats['positive_slices'] += int(gt_pos)
                    stats['empty_slices'] += int(not gt_pos)
                    stats['tp'] += int(gt_pos and pred_pos)
                    stats['fp'] += int((not gt_pos) and pred_pos)
                    stats['tn'] += int((not gt_pos) and (not pred_pos))
                    stats['fn'] += int(gt_pos and (not pred_pos))
                    stats['pixel_tp'] += intersection_pixels
                    stats['pixel_fp'] += pixel_fp
                    stats['pixel_tn'] += pixel_tn
                    stats['pixel_fn'] += pixel_fn

                case_stats[case_key]['total_slices'] += 1
                case_stats[case_key]['positive_slices'] += int(gt_pos)
                case_stats[case_key]['empty_slices'] += int(not gt_pos)
                case_stats[case_key]['pixel_tp'] += intersection_pixels
                case_stats[case_key]['pixel_fp'] += pixel_fp
                case_stats[case_key]['pixel_fn'] += pixel_fn
                case_stats[case_key]['pred_volume_slices'].append(pred_bool.copy())
                case_stats[case_key]['gt_volume_slices'].append(gt_bool.copy())

                dice_value = ''
                jaccard_value = ''
                hd95_value = ''
                asd_value = ''
                if gt_pos:
                    if pred_pos:
                        dice_value = _binary_dice(pred_bool, gt_bool)
                        jaccard_value = _binary_jaccard(pred_bool, gt_bool)
                        hd95_value = float(binary.hd95(pred_bool, gt_bool))
                        asd_value = float(binary.asd(pred_bool, gt_bool))
                    else:
                        dice_value = 0.0
                        jaccard_value = 0.0
                        hd95_value = 100.0
                        asd_value = 100.0

                    for stats in (domain_stats, overall_stats):
                        stats['positive_pred_empty_slices'] += int(not pred_pos)
                        stats['positive_dice_sum'] += float(dice_value)
                        stats['positive_jaccard_sum'] += float(jaccard_value)
                        stats['positive_hd95_sum'] += float(hd95_value)
                        stats['positive_asd_sum'] += float(asd_value)

                    case_stats[case_key]['positive_dice_sum'] += float(dice_value)
                    case_stats[case_key]['positive_jaccard_sum'] += float(jaccard_value)
                    case_stats[case_key]['positive_hd95_sum'] += float(hd95_value)
                    case_stats[case_key]['positive_asd_sum'] += float(asd_value)
                else:
                    for stats in (domain_stats, overall_stats):
                        stats['empty_pred_positive_slices'] += int(pred_pos)
                        stats['pred_pixels_on_empty_total'] += pred_pixels
                        stats['pred_pixels_on_empty_max'] = max(stats['pred_pixels_on_empty_max'], pred_pixels)

                if args.eval and args.save_img:
                    out_dir = os.path.join('./Output', args.dataset)
                    os.makedirs(out_dir, exist_ok=True)
                    score = dice_value if gt_pos else (1.0 if not pred_pos else 0.0)
                    save_path = os.path.join(
                        out_dir,
                        '{}_{}_{}.png'.format(domain_code, os.path.splitext(img_name)[0], round(100 * score, 2))
                    )
                    util.save_binary_mask(pred_onehot[j].cpu(), save_path)

                if gt_pos and pred_pos:
                    slice_confusion = 'TP'
                elif gt_pos and not pred_pos:
                    slice_confusion = 'FN'
                elif (not gt_pos) and pred_pos:
                    slice_confusion = 'FP'
                else:
                    slice_confusion = 'TN'

                slice_rows.append({
                    'domain': 'Dom{}'.format(domain_code),
                    'case_id': case_id,
                    'img_name': img_name,
                    'gt_label': 'non_empty' if gt_pos else 'empty',
                    'gt_positive': int(gt_pos),
                    'pred_positive': int(pred_pos),
                    'slice_confusion': slice_confusion,
                    'gt_pixels': gt_pixels,
                    'pred_pixels': pred_pixels,
                    'intersection_pixels': intersection_pixels,
                    'dice_on_positive_gt': dice_value,
                    'jaccard_on_positive_gt': jaccard_value,
                    'hd95_on_positive_gt': hd95_value,
                    'asd_on_positive_gt': asd_value,
                })

        domain_name = 'Dom{}'.format(domain_code)
        domain_case_results = [
            _finalize_case_stats(case_stats[key])
            for key in sorted(case_stats)
            if case_stats[key]['domain'] == domain_name
        ]
        domain_case_mean = _casewise_volume_mean(domain_case_results)

        finalized_domain = _finalize_lesion_stats(domain_stats)
        finalized_domain.update(domain_case_mean)
        finalized_domain['dice_policy'] = 'case-wise volume DC/JC/HD95/ASD averaged over cases in this domain'
        domain_results.append(finalized_domain)

        logging.info(
            'domain{} epoch {} :\n\t'
            'case_mean volume_dc / volume_jc / volume_hd95 / volume_asd: {:.2f} / {:.2f} / {:.2f} / {:.2f} | volume_cases: {}\n\t'
            'positive-slice analysis dc / jc / hd95 / asd: {:.2f} / {:.2f} / {:.2f} / {:.2f}\n\t'
            'slices pos/empty: {}/{}; TP/FP/TN/FN: {}/{}/{}/{}; '
            'sens/spec/prec/f1: {:.2f}/{:.2f}/{:.2f}/{:.2f}'.format(
                domain_code,
                epoch,
                100.0 * finalized_domain['volume_dice'],
                100.0 * finalized_domain['volume_jaccard'],
                finalized_domain['volume_hd95'],
                finalized_domain['volume_asd'],
                finalized_domain['valid_cases'],
                100.0 * finalized_domain['positive_dice'],
                100.0 * finalized_domain['positive_jaccard'],
                finalized_domain['positive_hd95'],
                finalized_domain['positive_asd'],
                finalized_domain['positive_slices'],
                finalized_domain['empty_slices'],
                finalized_domain['tp'],
                finalized_domain['fp'],
                finalized_domain['tn'],
                finalized_domain['fn'],
                100.0 * finalized_domain['slice_sensitivity'],
                100.0 * finalized_domain['slice_specificity'],
                100.0 * finalized_domain['slice_precision'],
                100.0 * finalized_domain['slice_f1'],
            )
        )

    all_slice_stats = _finalize_lesion_stats(overall_stats)
    case_results = [_finalize_case_stats(case_stats[key]) for key in sorted(case_stats)]
    domain_count = len(domain_results)
    domain_mean = {
        # Equal-weight average over domains. Each domain value is already case-wise.
        'volume_dice': _safe_div(sum(item['volume_dice'] for item in domain_results), domain_count),
        'volume_jaccard': _safe_div(sum(item['volume_jaccard'] for item in domain_results), domain_count),
        'volume_hd95': _safe_div(sum(item['volume_hd95'] for item in domain_results), domain_count),
        'volume_asd': _safe_div(sum(item['volume_asd'] for item in domain_results), domain_count),
        'positive_dice': _safe_div(sum(item['positive_dice'] for item in domain_results), domain_count),
        'positive_jaccard': _safe_div(sum(item['positive_jaccard'] for item in domain_results), domain_count),
        'positive_hd95': _safe_div(sum(item['positive_hd95'] for item in domain_results), domain_count),
        'positive_asd': _safe_div(sum(item['positive_asd'] for item in domain_results), domain_count),
    }
    case_mean = _casewise_volume_mean(case_results)
    overall_case_mean = _casewise_overall_metric_mean(case_results)

    logging.info(
        'epoch {} :\n\t'
        'main case_mean volume_dc / volume_jc / volume_hd95 / volume_asd: {:.2f} / {:.2f} / {:.2f} / {:.2f} | volume_cases: {}\n\t'
        'domain_mean casewise volume_dc / volume_jc / volume_hd95 / volume_asd: {:.2f} / {:.2f} / {:.2f} / {:.2f}\n\t'
        'positive-slice analysis domain_mean dc / jc / hd95 / asd: {:.2f} / {:.2f} / {:.2f} / {:.2f}\n\t'
        'all slices pos/empty: {}/{}; TP/FP/TN/FN: {}/{}/{}/{}; '
        'acc/sens/spec/prec/f1: {:.2f}/{:.2f}/{:.2f}/{:.2f}/{:.2f}\n\t'
        'empty FP rate: {:.2f}; pred pixels on empty total/mean/max: {}/{:.2f}/{}'.format(
            epoch,
            100.0 * case_mean['volume_dice'],
            100.0 * case_mean['volume_jaccard'],
            case_mean['volume_hd95'],
            case_mean['volume_asd'],
            case_mean['valid_cases'],
            100.0 * domain_mean['volume_dice'],
            100.0 * domain_mean['volume_jaccard'],
            domain_mean['volume_hd95'],
            domain_mean['volume_asd'],
            100.0 * domain_mean['positive_dice'],
            100.0 * domain_mean['positive_jaccard'],
            domain_mean['positive_hd95'],
            domain_mean['positive_asd'],
            all_slice_stats['positive_slices'],
            all_slice_stats['empty_slices'],
            all_slice_stats['tp'],
            all_slice_stats['fp'],
            all_slice_stats['tn'],
            all_slice_stats['fn'],
            100.0 * all_slice_stats['slice_accuracy'],
            100.0 * all_slice_stats['slice_sensitivity'],
            100.0 * all_slice_stats['slice_specificity'],
            100.0 * all_slice_stats['slice_precision'],
            100.0 * all_slice_stats['slice_f1'],
            100.0 * all_slice_stats['empty_false_positive_rate'],
            all_slice_stats['pred_pixels_on_empty_total'],
            all_slice_stats['pred_pixels_on_empty_mean'],
            all_slice_stats['pred_pixels_on_empty_max'],
        )
    )

    metrics_root = globals().get('snapshot_path', '.') or '.'
    os.makedirs(metrics_root, exist_ok=True)
    json_path = os.path.join(metrics_root, 'test_metrics_epoch_{}.json'.format(epoch))
    csv_path = os.path.join(metrics_root, 'test_slice_metrics_epoch_{}.csv'.format(epoch))
    metrics_out = {
        'dataset': args.dataset,
        'data_root': globals().get('train_data_path', args.data_root),
        'epoch': epoch,
        'dice_policy': (
            'Volume DC/JC/HD95/ASD are never computed by pooling pixels across a domain/all data. '
            'Each case is finalized first, then averaged. '
            'domain_mean is an equal-weight average of domain case-wise means. '
            'The Excel Dice(+ve) column is removed. Final confusion and classification metrics are computed over all slices combined.'
        ),
        'main_metric': {
            'name': 'case_mean_volume_dice',
            'value': case_mean['volume_dice'],
        },
        'domain_mean': domain_mean,
        'case_mean': case_mean,
        'overall_case_mean': overall_case_mean,
        'all_slice_detection': all_slice_stats,
        'domains': domain_results,
        'cases': case_results,
    }
    with open(json_path, 'w') as f:
        json.dump(metrics_out, f, indent=2)
    with open(csv_path, 'w', newline='') as f:
        fieldnames = [
            'domain',
            'case_id',
            'img_name',
            'gt_label',
            'gt_positive',
            'pred_positive',
            'slice_confusion',
            'gt_pixels',
            'pred_pixels',
            'intersection_pixels',
            'dice_on_positive_gt',
            'jaccard_on_positive_gt',
            'hd95_on_positive_gt',
            'asd_on_positive_gt',
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(slice_rows)
    logging.info('saved lesion test metrics: {} and {}'.format(json_path, csv_path))

    _save_lesion_summary_excel(
        args=args,
        epoch=epoch,
        domain_results=domain_results,
        domain_mean=domain_mean,
        overall_case_mean=overall_case_mean,
        all_slice_stats=all_slice_stats,
    )

    model.train()
    return (
        [case_mean['volume_dice']],
        [case_mean['volume_dice']],
        [case_mean['volume_jaccard']],
        [domain_mean['volume_hd95']],
        [domain_mean['volume_asd']],
    )


@torch.no_grad()
def test(args, model, test_dataloader, epoch):
    model.eval()
    if args.dataset == 'prostate_lesion':
        return _test_prostate_lesion_with_empty_metrics(args, model, test_dataloader, epoch)

    val_dice = [0.0] * n_part
    val_dc, val_jc, val_hd, val_asd = [0.0] * n_part, [0.0] * n_part, [0.0] * n_part, [0.0] * n_part
    domain_num = len(test_dataloader)
    num = 0
    for i in range(domain_num):
        cur_dataloader = test_dataloader[i]
        domain_val_dice = [0.0] * n_part
        domain_val_dc, domain_val_jc, domain_val_hd, domain_val_asd = [0.0] * n_part, [0.0] * n_part, [0.0] * n_part, [0.0] * n_part
        domain_code = i+1
        for batch_num,sample in enumerate(cur_dataloader):
            data = sample['image'].cuda()
            mask = sample['label'].cuda()
            if args.dataset == 'fundus':
                cup_mask = mask.eq(0).float()
                disc_mask = mask.le(128).float()
                mask = torch.cat((cup_mask.unsqueeze(1), disc_mask.unsqueeze(1)),dim=1)
            elif args.dataset == 'prostate':
                mask = mask.eq(0).long()
            elif args.dataset == 'prostate_lesion':
                mask = mask.gt(0).long()
            elif args.dataset == 'MNMS':
                mask_ = mask[:,...,0].eq(255).float()
                mask_[mask[:,...,1].eq(255)] = 2
                mask_[mask[:,...,2].eq(255)] = 3
                mask = mask_.long()
                
            output = model(data)
            mask = mask.cpu()
            output = output.cpu()
            if args.dataset == 'fundus':
                pred_label = torch.sigmoid(output).ge(0.5)
                pred_onehot = pred_label.clone()
                mask_onehot = mask.clone()
            elif args.dataset in ['prostate', 'prostate_lesion']:
                pred_label = torch.max(torch.softmax(output, dim=1), dim=1)[1]
                if args.dataset == 'prostate_lesion' and (
                    args.post_min_area > 0 or args.post_topk > 0 or args.post_fill_holes
                ):
                    pred_np = util.postprocess_binary_batch(
                        pred_label.numpy(),
                        min_area=args.post_min_area,
                        topk=args.post_topk,
                        fill_holes=bool(args.post_fill_holes),
                    )
                    pred_label = torch.from_numpy(pred_np).long()
                pred_onehot = pred_label.clone().unsqueeze(1)
                mask_onehot = mask.clone().unsqueeze(1)
            elif args.dataset == 'MNMS':
                pred_label = torch.max(torch.softmax(output, dim=1), dim=1)[1]
                pred_onehot = to_3d(pred_label)
                mask_onehot = to_3d(mask)
                
            dice = dice_calcu[args.dataset](np.asarray(pred_label),mask)
            avg_dice = sum(dice)/len(dice)
            
            if args.eval and args.save_img:
                out_dir = os.path.join('./Output', args.dataset)
                os.makedirs(out_dir, exist_ok=True)

                for j in range(len(data)):
                    img_name = os.path.splitext(sample['img_name'][j])[0]
                    save_path = os.path.join(
                        out_dir,
                        f"{domain_code}_{img_name}_{round(100 * avg_dice, 2)}.png"
                    )
                    util.save_binary_mask(pred_onehot[j].cpu(), save_path)
                    
            dc, jc, hd, asd = [0.0] * n_part, [0.0] * n_part, [0.0] * n_part, [0.0] * n_part
            for j in range(len(data)):
                for i, p in enumerate(part):
                    dc[i] += binary.dc(np.asarray(pred_onehot[j,i], dtype=bool),
                                            np.asarray(mask_onehot[j,i], dtype=bool))
                    jc[i] += binary.jc(np.asarray(pred_onehot[j,i], dtype=bool),
                                            np.asarray(mask_onehot[j,i], dtype=bool))
                    if pred_onehot[j,i].float().sum() < 1e-4:
                        hd[i] += 100
                        asd[i] += 100
                    else:
                        hd[i] += binary.hd95(np.asarray(pred_onehot[j,i], dtype=bool),
                                            np.asarray(mask_onehot[j,i], dtype=bool))
                        asd[i] += binary.asd(np.asarray(pred_onehot[j,i], dtype=bool),
                                            np.asarray(mask_onehot[j,i], dtype=bool))
            for i, p in enumerate(part):
                dc[i] /= len(data)
                jc[i] /= len(data)
                hd[i] /= len(data)
                asd[i] /= len(data)
            for i in range(len(domain_val_dice)):
                domain_val_dice[i] += dice[i]
                domain_val_dc[i] += dc[i]
                domain_val_jc[i] += jc[i]
                domain_val_hd[i] += hd[i]
                domain_val_asd[i] += asd[i]
                
        for i in range(len(domain_val_dice)):
            domain_val_dice[i] /= len(cur_dataloader)
            val_dice[i] += domain_val_dice[i]
            domain_val_dc[i] /= len(cur_dataloader)
            val_dc[i] += domain_val_dc[i]
            domain_val_jc[i] /= len(cur_dataloader)
            val_jc[i] += domain_val_jc[i]
            domain_val_hd[i] /= len(cur_dataloader)
            val_hd[i] += domain_val_hd[i]
            domain_val_asd[i] /= len(cur_dataloader)
            val_asd[i] += domain_val_asd[i]

        # --- REPLACE THIS DOMAIN-WISE LOGGING BLOCK ---
        label = " / ".join([f"val_{p}_dc" for p in part]) + ":  "
        dc_vals_pct = [round(100.0 * domain_val_dc[n], 2) for n in range(len(part))]
        vals = " / ".join([f"{v:.2f}" for v in dc_vals_pct])
        text = f"domain{domain_code} epoch {epoch} :\n\t{label}{vals}"
        logging.info(text)
        # --- END REPLACEMENT ---

        
    model.train()
    for i in range(len(val_dice)):
        val_dice[i] /= domain_num
        val_dc[i] /= domain_num
        val_jc[i] /= domain_num
        val_hd[i] /= domain_num
        val_asd[i] /= domain_num
    # --- REPLACE THIS EPOCH-LEVEL LOGGING BLOCK ---
    # average across parts (arrays already averaged across domains above)
    avg_dc  = sum(val_dc)  / len(val_dc)   if len(val_dc)  > 0 else 0.0
    avg_jc  = sum(val_jc)  / len(val_jc)   if len(val_jc)  > 0 else 0.0
    avg_hd  = sum(val_hd)  / len(val_hd)   if len(val_hd)  > 0 else 0.0
    avg_asd = sum(val_asd) / len(val_asd)  if len(val_asd) > 0 else 0.0

    # dc and jc in %, hd and asd as-is; all rounded to 2
    text = (
        f"epoch {epoch} :\n\t"
        f"avg_dc / avg_jc / avg_hd / avg_asd: "
        f"{avg_dc*100.0:.2f} / {avg_jc*100.0:.2f} / {avg_hd:.2f} / {avg_asd:.2f}"
    )
    logging.info(text)
    # --- END REPLACEMENT ---

    return val_dice, val_dc, val_jc, val_hd, val_asd
    
def main(args, snapshot_path):

    if args.dataset == 'fundus':
        num_channels = 3
        patch_size = 256
        num_classes = 2
        if args.domain_num >= 4:
            args.domain_num = 4
    elif args.dataset == 'prostate':
        num_channels = 1
        patch_size = 384
        num_classes = 2
        if args.domain_num >= 6:
            args.domain_num = 6
    elif args.dataset == 'prostate_lesion':
        lesion_modalities = split_csv(args.lesion_modalities) or ['t2w', 'adc']
        num_channels = len(lesion_modalities) + int(args.add_adc_sobel)
        patch_size = 224
        num_classes = 2
        if args.domain_num >= 3:
            args.domain_num = 3
    elif args.dataset == 'MNMS':
        num_channels = 1
        patch_size = 288
        num_classes = 4
        if args.domain_num >= 4:
            args.domain_num = 4
    if args.dataset == 'prostate_lesion':
        lesion_modalities = split_csv(args.lesion_modalities) or ['t2w', 'adc']
        normal_toTensor = transforms.Compose([
            tr.LesionNormalize(
                modalities=lesion_modalities,
                mode=args.lesion_norm,
                add_adc_sobel=bool(args.add_adc_sobel),
            ),
            tr.ToTensor()
        ])
    else:
        normal_toTensor = transforms.Compose([
            tr.Normalize_tf(),
            tr.ToTensor()
        ])

    domain_num = args.domain_num
    test_dataset = []
    test_dataloader = []
    dataset_kwargs = {}
    if args.dataset == 'prostate_lesion':
        dataset_kwargs = {
            'modalities': args.lesion_modalities,
            'modality_dirs': args.lesion_modality_dirs,
        }
    for i in range(1, domain_num+1):
        cur_dataset = dataset(base_dir=train_data_path, phase='test', splitid=-1, domain=[i], normal_toTensor=normal_toTensor,
                              **dataset_kwargs)
        test_dataset.append(cur_dataset)
    for i in range(0,domain_num):
        cur_dataloader = DataLoader(test_dataset[i], batch_size = args.test_bs, shuffle=False, num_workers=0, pin_memory=True)
        test_dataloader.append(cur_dataloader)

    def create_model(ema=False):
        model = build_segmentation_model(
            args.model,
            n_channels=num_channels,
            n_classes=num_classes,
            patch_size=patch_size,
            encoder_name=args.encoder_name,
            encoder_weights=args.encoder_weights,
        )
        if ema:
            for param in model.parameters():
                param.detach_()
        return model.cuda()

    model = create_model()

    if args.eval:
        args.lb_domain = 1
        best_model_path = os.path.join(snapshot_path, '{}_avg_dice_best_model.pth'.format(args.model))
        model.load_state_dict(torch.load(best_model_path, map_location='cpu'))
        test(args, model,test_dataloader,args.lb_domain)
        exit()


if __name__ == "__main__":
    repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    snapshot_path = os.path.join(repo_root, "model", args.dataset, args.save_name) + "/"
    default_data_roots = {
        'fundus': os.path.join(repo_root, 'data', 'Fundus'),
        'prostate': os.path.join(repo_root, 'data', 'Prostate'),
        'prostate_lesion': os.path.join(repo_root, 'data', 'Prostate_Lesion'),
        'MNMS': os.path.join(repo_root, 'data', 'MNMS'),
    }
    train_data_path = args.data_root if args.data_root else default_data_roots[args.dataset]

    os.environ['CUDA_VISIBLE_DEVICES'] = args.gpu

    if not os.path.exists(snapshot_path):
        raise FileNotFoundError('Model folder not found for testing: {}'.format(snapshot_path))
    logging.basicConfig(filename=snapshot_path + "/log.txt", level=logging.INFO,
                        format='[%(asctime)s.%(msecs)03d] %(message)s', datefmt='%H:%M:%S')
    logging.getLogger().addHandler(logging.StreamHandler(sys.stdout))
    cmd = " ".join(["python"] + sys.argv)
    logging.info(cmd)
    logging.info(str(args))

    main(args, snapshot_path)
